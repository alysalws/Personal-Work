from selenium import webdriver;import time
import numpy as np
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium import webdriver

import os
import time
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from openpyxl import Workbook
from openpyxl import load_workbook

from selenium.webdriver.support.ui import Select
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
import os
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
wb = Workbook()
ws = wb.active
ws.append(['Date', 'Author','Pros','Cons'])

user_agents_file = open("C:/Users/Alysa/Documents/UCL/UCL/2- MSIN0166 Data Engineering/Individual project/user-agents.txt", 'r')
useragents = list(user_agents_file.readlines())
us_ag = np.random.choice(useragents).replace("\n", '')
headers = {
'accept-encoding': 'gzip, deflate, br',
'accept-language': 'en-US,en;q=0.9',
'referer': 'https://www.glassdoor.com/',
'user-agent': us_ag,
'content-type': 'application/json'
}

options = Options()
options.add_argument("--window-size=1920,1080")
options.add_argument(f'headers={headers}')

path = os.path.abspath("C:/Users/Alysa/chromedriver_win32 (1)/chromedriver.exe")
driver = webdriver.Chrome(executable_path=path, options=options)
driver.get('https://www.glassdoor.co.uk/profile/login_input.htm?%27')

id = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//input[@id="inlineUserEmail"]')))
time.sleep(1)
# Enter Username below:
id.send_keys("  ") #input the email which you signed up for glassdoor
password = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//input[@id="inlineUserPassword"]')))
#Enter Password below:
password.send_keys("  ") # input your own password
password.submit()
time.sleep(1)

btn_company = driver.find_element(By.XPATH,'//a[@data-test="site-header-companies"]').click()
time.sleep(1)
input_city = driver.find_element(By.XPATH,'//input[@data-test="search-bar-location-input"]')
input_city.send_keys(".")
time.sleep(1)
input_bank = driver.find_element(By.XPATH,'//input[@data-test="search-bar-keyword-input"]')
input_bank.send_keys("Citigroup Inc")
time.sleep(1)
btn_search = driver.find_element(By.XPATH,'//button[@data-test="search-bar-submit"]').click()
time.sleep(1)
click_review = driver.find_element(By.XPATH,'//a[@data-label="Reviews"]').click()
time.sleep(1)

all_links = []
link_no = 0

while link_no < 6:
    link_no += 1
    time.sleep(2)
    cards = WebDriverWait(driver,20).until(EC.presence_of_all_elements_located((By.XPATH,'//div[@class="gdReview"]')))
    for card in cards:
        authorinfo = card.find_element(By.CSS_SELECTOR,'span[class="authorInfo"]').text
        date = authorinfo.split(" - ")[0]
        author = authorinfo.split(" - ")[1]
        pros = card.find_element(By.CSS_SELECTOR,'span[data-test="pros"]').text
        cons = card.find_element(By.CSS_SELECTOR,'span[data-test="cons"]').text
        this_row = [date, author, pros, cons]
        ws.append(this_row)
        wb.save("glassdoorData.xlsx")
    time.sleep(1)

    while True:
        time.sleep(1)
        try:
            next_btn = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//button[@data-test="pagination-next"]')))
            time.sleep(1)
            try:
                next_btn.click()
                time.sleep(1)
                break
            except Exception as e:
                pass
        except Exception as e:
            pass
#
